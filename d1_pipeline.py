import cv2
import numpy as np
from skimage import measure
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

IMAGE_DIR       = Path("D1 Photos (STP + WAS Only)")
XLSX_IN         = "S26 First deployment of sticky traps.xlsx"
OUTPUT_DIR      = Path("outputs")
INSECT_MIN_AREA = 20

OUTPUT_DIR.mkdir(exist_ok=True)

def parse_filename(stem):
    parts = stem.strip().split()
    if len(parts) != 2: return None
    site_plot, dep_wind = parts
    site_code = site_plot[0]
    try: plot_num = int(site_plot[1:])
    except: return None
    return {
        "site_code":  site_code,
        "site":       "Saint Paul" if site_code == "S" else "Waseca",
        "plot":       plot_num,
        "deployment": dep_wind[:2],
        "wind_dir":   dep_wind[2:],
    }

def analyze(img_path):
    img_bgr = cv2.imread(str(img_path))
    if img_bgr is None: return None

    img_bgr = cv2.resize(img_bgr, (0,0), fx=0.25, fy=0.25)

    # Crop bottom 15% to remove label/handwriting
    img_bgr = img_bgr[:int(img_bgr.shape[0] * 0.85), :]

    img_hsv = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2HSV)
    h, s, v = img_hsv[:,:,0], img_hsv[:,:,1], img_hsv[:,:,2]

    # ── TRAP MASK ──────────────────────────────────────────────────────────────
    # Use strict (high saturation) yellow to find trap outline
    is_yellow_strict = (h>=15)&(h<=45)&(s>180)&(v>80)
    kernel = np.ones((15,15), np.uint8)
    yellow_clean = cv2.morphologyEx(is_yellow_strict.astype(np.uint8), cv2.MORPH_CLOSE, kernel)
    yellow_clean = cv2.morphologyEx(yellow_clean, cv2.MORPH_OPEN, kernel)
    contours, _ = cv2.findContours(yellow_clean, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not contours: return None

    largest = max(contours, key=cv2.contourArea)
    trap_mask = np.zeros(img_bgr.shape[:2], dtype=np.uint8)
    cv2.drawContours(trap_mask, [largest], -1, 255, thickness=cv2.FILLED)
    trap_mask = trap_mask.astype(bool)
    total = trap_mask.sum()
    if total == 0: return None

    img_rgb = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)

    # ── FOREGROUND DETECTION ───────────────────────────────────────────────────
    # Only exclude very high saturation yellow (pure trap background)
    is_yel_strict = (h>=15)&(h<=45)&(s>180)&(v>80)
    is_glare      = v > 230

    # Dark pixels = dark insects (flies, beetles)
    is_dark = v < 180

    # Semi-transparent wing pixels: yellow hue but LOWER saturation than background
    # These are light grey gnats whose wings pick up the yellow background colour
    is_wing = (h>=15)&(h<=45)&(s>=50)&(s<=180)&(v>60)&(v<230)

    foreground = (is_dark | is_wing) & ~is_yel_strict & ~is_glare & trap_mask

    kernel2 = np.ones((2,2), np.uint8)
    cleaned = cv2.morphologyEx(foreground.astype(np.uint8), cv2.MORPH_OPEN, kernel2)
    labeled = measure.label(cleaned, connectivity=2)
    regions = measure.regionprops(labeled)

    insect_mask   = np.zeros_like(cleaned, dtype=bool)
    sediment_mask = np.zeros_like(cleaned, dtype=bool)
    for r in regions:
        if r.area < 5: continue
        if r.area >= INSECT_MIN_AREA: insect_mask[labeled==r.label] = True
        else: sediment_mask[labeled==r.label] = True

    return {
        "insect_pct":    round(100 * insect_mask.sum()   / total, 2),
        "sediment_pct":  round(100 * sediment_mask.sum() / total, 2),
        "img_rgb":       img_rgb,
        "insect_mask":   insect_mask,
        "sediment_mask": sediment_mask,
        "trap_mask":     trap_mask,
    }

# ── LOAD DATASHEET ─────────────────────────────────────────────────────────────
df_meta = pd.read_excel(XLSX_IN, sheet_name="Plots")
df_meta["Plot"] = df_meta["Plot"].astype(int)
df_meta["Wind_direction"] = df_meta["Wind_direction"].str.strip()

images = sorted(IMAGE_DIR.glob("*.jpg"))
print(f"Found {len(images)} images\n")

rows = []
for img_path in images:
    info = parse_filename(img_path.stem)
    if not info:
        print(f"  SKIP: {img_path.name}")
        continue
    data = analyze(img_path)
    if not data:
        print(f"  FAIL: {img_path.name}")
        continue

    print(f"  {img_path.stem}: insects={data['insect_pct']}%  sediment={data['sediment_pct']}%")

    match = df_meta[
        (df_meta["Site"].str.contains(info["site"], case=False)) &
        (df_meta["Plot"] == info["plot"]) &
        (df_meta["Wind_direction"].str.upper() == info["wind_dir"].upper())
    ]
    trt = trt_rate = trt_type = rep = "N/A"
    if not match.empty:
        row = match.iloc[0]
        trt=row["Trt"]; trt_rate=row["Trt_rate"]
        trt_type=row["Trt_type"]; rep=row["Rep"]

    rows.append({
        "filename": img_path.stem, "site": info["site"], "plot": info["plot"],
        "rep": rep, "deployment": info["deployment"], "wind_dir": info["wind_dir"],
        "treatment": trt, "trt_rate": trt_rate, "trt_type": trt_type,
        "insect_pct": data["insect_pct"], "sediment_pct": data["sediment_pct"],
    })

    # Save result image
    overlay = data["img_rgb"].copy()
    overlay[data["insect_mask"]]   = [220, 50,  50]
    overlay[data["sediment_mask"]] = [50,  100, 220]
    overlay[~data["trap_mask"]]    = [40,  40,  40]

    fig, axes = plt.subplots(1, 2, figsize=(13, 6))
    axes[0].imshow(data["img_rgb"]); axes[0].axis("off")
    axes[0].set_title(f"Original — {img_path.stem}", fontsize=9)
    axes[1].imshow(overlay); axes[1].axis("off")
    axes[1].set_title(
        f"Insect: {data['insect_pct']}%  |  Sediment: {data['sediment_pct']}%\n"
        f"Site: {info['site']}  Plot: {info['plot']}  Wind: {info['wind_dir']}  Trt: {trt}",
        fontsize=9)
    ip = mpatches.Patch(color=(220/255,50/255,50/255),  label=f"Insect ({data['insect_pct']}%)")
    sp = mpatches.Patch(color=(50/255,100/255,220/255), label=f"Sediment ({data['sediment_pct']}%)")
    axes[1].legend(handles=[ip,sp], loc="lower right", fontsize=8)
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / f"{img_path.stem}_result.png", dpi=100, bbox_inches="tight")
    plt.close()

# ── BUILD EXCEL ────────────────────────────────────────────────────────────────
df = pd.DataFrame(rows)
wb = Workbook()
FOREST="2C5F2D"; WHITE="FFFFFF"; LIGHT="F0F4F0"

ws1 = wb.active; ws1.title = "Full Results"
h1 = ["Filename","Site","Plot","Rep","Deployment","Wind Dir",
      "Treatment","Trt Rate","Trt Type","Insect Area %","Sediment Area %"]
ws1.append(h1)
for c in range(1,len(h1)+1):
    cell=ws1.cell(row=1,column=c)
    cell.font=Font(bold=True,color=WHITE,name="Arial")
    cell.fill=PatternFill("solid",start_color=FOREST)
    cell.alignment=Alignment(horizontal="center")
for r in rows:
    ws1.append([r["filename"],r["site"],r["plot"],r["rep"],r["deployment"],
                r["wind_dir"],r["treatment"],r["trt_rate"],r["trt_type"],
                r["insect_pct"],r["sediment_pct"]])
for i in range(2,len(rows)+2):
    if i%2==0:
        for j in range(1,len(h1)+1):
            ws1.cell(row=i,column=j).fill=PatternFill("solid",start_color=LIGHT)
for i,w in enumerate([22,12,6,5,12,9,12,10,12,14,16],1):
    ws1.column_dimensions[get_column_letter(i)].width=w

ws2 = wb.create_sheet("Summary by Treatment")
summary = df.groupby(["treatment","trt_rate","trt_type"]).agg(
    n_images=("filename","count"), avg_insect=("insect_pct","mean"),
    avg_sediment=("sediment_pct","mean"), max_sediment=("sediment_pct","max"),
    min_sediment=("sediment_pct","min")).reset_index().round(3)
h2=["Treatment","Rate","Type","N Images","Avg Insect %","Avg Sediment %","Max Sediment %","Min Sediment %"]
ws2.append(h2)
for c in range(1,len(h2)+1):
    cell=ws2.cell(row=1,column=c)
    cell.font=Font(bold=True,color=WHITE,name="Arial")
    cell.fill=PatternFill("solid",start_color=FOREST)
    cell.alignment=Alignment(horizontal="center")
for _,r in summary.iterrows():
    ws2.append([r["treatment"],r["trt_rate"],r["trt_type"],r["n_images"],
                r["avg_insect"],r["avg_sediment"],r["max_sediment"],r["min_sediment"]])
for i,w in enumerate([12,8,12,10,14,16,16,16],1):
    ws2.column_dimensions[get_column_letter(i)].width=w

ws3 = wb.create_sheet("Summary by Site")
site_sum = df.groupby(["site","treatment"]).agg(
    n_images=("filename","count"), avg_insect=("insect_pct","mean"),
    avg_sediment=("sediment_pct","mean")).reset_index().round(3)
h3=["Site","Treatment","N Images","Avg Insect %","Avg Sediment %"]
ws3.append(h3)
for c in range(1,len(h3)+1):
    cell=ws3.cell(row=1,column=c)
    cell.font=Font(bold=True,color=WHITE,name="Arial")
    cell.fill=PatternFill("solid",start_color=FOREST)
    cell.alignment=Alignment(horizontal="center")
for _,r in site_sum.iterrows():
    ws3.append([r["site"],r["treatment"],r["n_images"],r["avg_insect"],r["avg_sediment"]])
for i,w in enumerate([14,12,10,14,16],1):
    ws3.column_dimensions[get_column_letter(i)].width=w

wb.save("outputs/D1_analysis_results.xlsx")
print(f"\nExcel saved!")
print("\n── SUMMARY BY TREATMENT ──────────────────────────────")
print(summary.to_string(index=False))
print("\n── SUMMARY BY SITE ───────────────────────────────────")
print(site_sum.to_string(index=False))
print(f"\nDone! {len(rows)} images analyzed.")